"""
金融研究报告图表工具函数
============================

提供符合研报规范的高度封装图表生成函数。

Usage:
    from chart_utils import create_research_chart, create_price_volume_chart

    wb = create_research_chart(
        df,
        chart_type="line",
        title="股价走势",
        x_col="date",
        y_cols="close",
        output_path="report.xlsx"
    )
"""

from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, Series
from openpyxl.chart.shapes import GraphicalProperties
import pandas as pd
from typing import Optional, List, Union


def create_research_chart(
    df: pd.DataFrame,
    chart_type: str = "line",
    title: str = "",
    x_col: str = None,
    y_cols: Union[str, List[str]] = None,
    y_axis_title: str = "",
    x_axis_title: str = "",
    y_format: str = "0.00",
    show_data_labels: bool = False,
    show_last_label_only: bool = False,
    add_mean_line: bool = False,
    color_up: str = "FF0000",
    color_down: str = "00B050",
    tick_skip: int = 5,
    height: int = 10,
    width: int = 20,
    output_path: str = None
) -> Workbook:
    """
    生成符合研报规范的 Excel 图表

    Parameters:
    -----------
    df : pd.DataFrame
        数据源，必须包含 x_col 和 y_cols 指定的列
    chart_type : str
        图表类型："line"(折线图), "bar"(柱状图)
    title : str
        图表标题
    x_col : str
        X轴数据列名（日期/类别）
    y_cols : str or List[str]
        Y轴数据列名，支持多序列
    y_axis_title : str
        Y轴标题（建议带单位，如"价格（元）"）
    x_axis_title : str
        X轴标题
    y_format : str
        Y轴数字格式，默认"0.00"，百分比用"0.00%"
    show_data_labels : bool
        是否显示所有数据标签
    show_last_label_only : bool
        是否仅显示最后一个数据点的标签（标注最新值）
    add_mean_line : bool
        是否添加均值参考线
    color_up : str
        上涨/正值颜色（默认红色 FF0000）
    color_down : str
        下跌/负值颜色（默认绿色 00B050）
    tick_skip : int
        X轴标签间隔（防止重叠）
    height : int
        图表高度（厘米）
    width : int
        图表宽度（厘米）
    output_path : str
        输出文件路径，None则返回Workbook对象

    Returns:
    --------
    Workbook : openpyxl Workbook对象

    Examples:
    ---------
    >>> # 股价走势图
    >>> df = pd.DataFrame({
    ...     'date': ['2024-01', '2024-02', '2024-03'],
    ...     'close': [10.5, 11.2, 10.8]
    ... })
    >>> wb = create_research_chart(
    ...     df, chart_type="line",
    ...     title="股价走势",
    ...     x_col="date", y_cols="close",
    ...     y_axis_title="价格（元）",
    ...     show_last_label_only=True,
    ...     output_path="股价走势.xlsx"
    ... )

    >>> # 营收利润对比图
    >>> wb = create_research_chart(
    ...     df, chart_type="bar",
    ...     title="营业收入与净利润",
    ...     x_col="period", y_cols=["revenue", "profit"],
    ...     y_axis_title="金额（亿元）",
    ...     y_format="0.0"
    ... )
    """
    # 标准化 y_cols
    if isinstance(y_cols, str):
        y_cols = [y_cols]

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    # 写入表头
    headers = [x_col] + y_cols
    ws.append(headers)

    # 写入数据
    for _, row in df.iterrows():
        ws.append([row[col] for col in headers])

    # 创建图表
    if chart_type == "line":
        chart = LineChart()
    elif chart_type == "bar":
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
    else:
        chart = LineChart()

    # 设置标题和轴
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title

    # 设置图表尺寸
    chart.height = height
    chart.width = width

    # 设置数据区域
    data_start_row = 1
    data_end_row = len(df) + 1

    for i, y_col in enumerate(y_cols):
        col_idx = headers.index(y_col) + 1
        data_ref = Reference(ws, min_col=col_idx, min_row=data_start_row,
                             max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=data_end_row)

        chart.add_data(data_ref, titles_from_data=True)
        if i == 0:
            chart.set_categories(cats_ref)

    # 设置Y轴格式
    chart.y_axis.numFmt = y_format

    # 设置X轴标签间隔
    chart.x_axis.tickLblSkip = tick_skip

    # 网格线设置（仅保留水平主网格线）
    chart.x_axis.majorGridlines = None

    # 数据标签设置
    if show_data_labels:
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

    # 仅显示最后一个标签
    if show_last_label_only and chart.series:
        series = chart.series[0]
        last_idx = len(df) - 1
        pt = DataPoint(idx=last_idx)
        pt.graphicalProperties = GraphicalProperties(solidFill=color_up)
        series.data_points = [pt]

    # 柱状图涨跌着色
    if chart_type == "bar" and len(y_cols) == 1:
        series = chart.series[0]
        values = df[y_cols[0]].tolist()
        for i, val in enumerate(values):
            pt = DataPoint(idx=i)
            fill_color = color_up if val >= 0 else color_down
            pt.graphicalProperties = GraphicalProperties(solidFill=fill_color)
            series.data_points.append(pt)

    # 添加均值参考线
    if add_mean_line and chart.series:
        series = chart.series[0]
        values = df[y_cols[0]].tolist()
        mean_val = sum(values) / len(values)

        # 在工作表添加均值列
        mean_col = len(headers) + 1
        ws.cell(row=1, column=mean_col, value="均值")
        for i in range(2, data_end_row + 1):
            ws.cell(row=i, column=mean_col, value=mean_val)

        mean_ref = Reference(ws, min_col=mean_col, min_row=1, max_row=data_end_row)
        mean_series = Series(mean_ref, title="均值")
        chart.series.append(mean_series)

    # 图例位置（多序列时显示在底部）
    if len(y_cols) > 1 or add_mean_line:
        chart.legend.position = "b"
    else:
        chart.legend = None

    # 添加图表到工作表
    ws.add_chart(chart, "E2")

    # 添加数据来源注脚
    ws["A" + str(data_end_row + 2)] = "数据来源：Tushare"

    # 保存或返回
    if output_path:
        wb.save(output_path)

    return wb


def create_price_volume_chart(
    df: pd.DataFrame,
    title: str = "股价与成交量",
    date_col: str = "date",
    price_col: str = "close",
    volume_col: str = "volume",
    output_path: str = None
) -> Workbook:
    """
    生成股价+成交量组合图表（双Y轴）

    Parameters:
    -----------
    df : pd.DataFrame
        包含日期、收盘价、成交量的数据
    title : str
        图表标题
    date_col, price_col, volume_col : str
        各数据列名
    output_path : str
        输出路径

    Returns:
    --------
    Workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    # 写入数据
    ws.append([date_col, price_col, volume_col])
    for _, row in df.iterrows():
        ws.append([row[date_col], row[price_col], row[volume_col]])

    # 创建价格折线图（主Y轴）
    price_chart = LineChart()
    price_chart.title = title
    price_chart.y_axis.title = "价格（元）"
    price_chart.x_axis.title = date_col

    price_ref = Reference(ws, min_col=2, min_row=1, max_row=len(df)+1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)

    price_chart.add_data(price_ref, titles_from_data=True)
    price_chart.set_categories(cats_ref)
    price_chart.y_axis.numFmt = "0.00"
    price_chart.x_axis.tickLblSkip = max(1, len(df) // 10)

    # 创建成交量柱状图（次Y轴）
    vol_chart = BarChart()
    vol_chart.type = "col"
    vol_chart.grouping = "clustered"
    vol_chart.y_axis.axId = 200
    vol_chart.y_axis.title = "成交量（手）"

    vol_ref = Reference(ws, min_col=3, min_row=1, max_row=len(df)+1)
    vol_chart.add_data(vol_ref, titles_from_data=True)

    # 组合图表
    price_chart += vol_chart
    price_chart.y_axis.crosses = "max"

    # 尺寸设置
    price_chart.height = 10
    price_chart.width = 20

    # 隐藏网格线
    price_chart.x_axis.majorGridlines = None

    ws.add_chart(price_chart, "E2")
    ws["A" + str(len(df) + 3)] = "数据来源：Tushare"

    if output_path:
        wb.save(output_path)

    return wb
